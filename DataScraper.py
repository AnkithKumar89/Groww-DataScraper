import os
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# Load the input Excel file
input_path = 'input.xlsx'
if os.path.isfile(input_path):
    input_file = load_workbook(input_path)
    input_sheet = input_file.active
else:
    print(f'Input file "{input_path}" does not exist.')
    exit()

# Define the headings for the output sheet
headings = ['mutual_fund_url', 'manager_name', 'join_date']

# Check if the output file exists
if os.path.isfile('output.xlsx'):
    # If the output file exists, load it
    output_file = load_workbook('output.xlsx')
    output_sheet = output_file.active
else:
    # If the output file doesn't exist, create a new one
    output_file = Workbook()
    output_sheet = output_file.active
output_sheet.append(headings)

#store the urls in a list
urls=[]
#Store the joining date in a list 
join_date_text=[]
#Store the name in a list
name_text=[]
count=0
# Iterate over each cell in column D starting from row 2
for cell_tuple in input_sheet['D2:D{}'.format(input_sheet.max_row)]:
    # Get the URL from the current cell value
    url = cell_tuple[0].value
    urls.append(url)
    count=count+1
    print(count)
    # Send a GET request to the web page
    response = requests.get(url)
    # Parse the HTML content of the page with BeautifulSoup
    soup = BeautifulSoup(response.content, 'html.parser')
    fund_managers=soup.find_all('div',{'class':'cur-po fm982AboutFundManager'})
    for fund_manager in fund_managers:
        fund_manager_name=fund_manager.find('div',{'class':'fm982CardText'})
        name=fund_manager_name.find('div')
        join_date=fund_manager_name.find('div').next_sibling
        span_date=fund_manager.find('span')
        span_date.extract()
        name_text.append(name.text)
        join_date_text.append(join_date.text.strip())
data=zip(urls,name_text,join_date_text)
    
# Iterate over each pair of values and write them to the output sheet
for row, (url, name, date) in enumerate(data, start=2):
    output_sheet.cell(row=row, column=1, value=url)
    output_sheet.cell(row=row, column=2, value=name)
    output_sheet.cell(row=row, column=3, value=date)

# Save the output file
output_file.save('output.xlsx')